import openpyxl
import pandas as pd
from openpyxl.workbook import Workbook


def pd_dtype_to_str(dtype):
    if pd.api.types.is_integer_dtype(dtype):
        return "int"
    if pd.api.types.is_float_dtype(dtype):
        return "float"
    if pd.api.types.is_bool_dtype(dtype):
        return "bool"
    if pd.api.types.is_datetime64_any_dtype(dtype):
        return "date"
    return "str"

def load_schema(template_path):
    wb = openpyxl.load_workbook(template_path, data_only=True, read_only=True)
    out = {"sheets": {}}
    for ws in wb.worksheets:
        df = pd.DataFrame(ws.values)
        if df.empty:
            continue
        header = df.iloc[0].tolist()

        # Clean the header
        cleaned_header = []
        for i, h in enumerate(header):
            col_name = str(h).strip() if h else f"Unnamed_{i}"
            cleaned_header.append(col_name)

        df.columns = cleaned_header
        df = df.drop(index=0).reset_index(drop=True)

        columns_info = [
            {"name": c, "type": pd_dtype_to_str(df[c].dtype)}
            for c in cleaned_header
        ]

        out["sheets"][ws.title] = {"columns": columns_info}
    return out

def json_to_xlsx(pred_json: dict, schema_json: dict, dest_path: str):
    """
    Build a new XLSX file from a JSON schema dict and prediction data.

    pred_json: {
      "<sheet_name>": [
        { "col1": val1, "col2": val2, … },
        …
      ],
      …
    }
    schema_json: {
      "sheets": {
        "<sheet_name>": {
          "columns": [
            { "name": "<col1>", "type": "<type1>" },
            { "name": "<col2>", "type": "<type2>" },
            …
          ]
        },
        …
      }
    }
    dest_path: where to write the .xlsx
    """
    wb = Workbook()
    # Remove the default sheet if we’ll be creating our own
    default = wb.active
    if default and default.title == "Sheet":
        wb.remove(default)

    for sheet_name, spec in schema_json["sheets"].items():
        ws = wb.create_sheet(title=sheet_name)
        col_order = [c["name"] for c in spec["columns"]]

        # Header row
        for col_idx, col_name in enumerate(col_order, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Data rows
        rows = pred_json.get(sheet_name, [])
        for r_idx, record in enumerate(rows, start=2):
            for c_idx, col_name in enumerate(col_order, start=1):
                ws.cell(row=r_idx, column=c_idx, value=record.get(col_name))

    # Save to disk
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
    wb.save(dest_path)
